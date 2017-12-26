# -*- coding:utf-8-*-

import os
import sys
import sqlite3
from openpyxl import Workbook
from sqlite3 import DatabaseError
from PyQt5 import QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QBrush, QFont
from PyQt5.QtWidgets import QApplication, QTableWidgetItem, QMessageBox
from PyQt5 import QtCore, QtGui, QtWidgets
from pyqt2nd import TableUi, InfoDlg, tableuires_rc

_author = "游侠最光阴"  # 设置作者
_version = "V0.9.8"  # 设置版本号


class TableUiOpt(QtWidgets.QMainWindow, TableUi.Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setupUi(self)
        self.init_table()
        self.init_db()

        # 菜单菜单列表
        self.addaction.triggered.connect(self.add_table_row)  # 添加一行
        self.deleteaction.triggered.connect(self.delete_table_row)  # 删除一行

        # 内容菜单列表
        self.additemaction.triggered.connect(self.insert_row)  # 编辑一行
        self.updateitemaction.triggered.connect(self.update_row)  # 更新一行
        self.deleteitemaction.triggered.connect(self.delete_row)  # 删除一行
        self.savetodbaction.triggered.connect(self.save_to_db)  # 将数据保存到数据库
        self.exporttoexcelaction.triggered.connect(self.save_to_excel)  # 将数据保存到excel文件

        # 说明菜单列表
        self.declareaction.triggered.connect(self.show_declaration)  # 显示声明
        self.qtaction.triggered.connect(self.about_qt)

    # 在此处去初始化table的行和列，而不去在UI转换的文件之中操作，这样方便操作
    def init_table(self):
        self.setWindowTitle("简易通讯录")
        # 将窗口设置成为固定大小的模式，没有放大和缩小的按钮
        height = self.tableWidget.height()
        width = self.tableWidget.width()
        self.setFixedHeight(height + 5)
        self.setFixedWidth(width + 5)

        self.tableWidget.setFixedWidth(width)
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setRowCount(12)
        mlabel = ['序号', '姓名', '年龄', '性别', '电话', '备注']
        self.tableWidget.setHorizontalHeaderLabels(mlabel)  # 设置表头
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)  # 设置双击不编辑
        self.tableWidget.setAlternatingRowColors(True)  # 改变正在修改的行的颜色
        # 如何给某一个cell设置tooltip?
        # 设置表格的颜色
        for index in range(self.tableWidget.columnCount()):
            headItem = self.tableWidget.horizontalHeaderItem(index)
            headItem.setFont(QFont("微软雅黑", 10, QFont.Bold))  # 设置字体
            headItem.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # 设置
            if (index % 2 == 0):
                headItem.setBackground(QBrush(Qt.blue))  # 设置背景，但好像没有用
            else:
                headItem.setForeground(QBrush(Qt.darkMagenta))
        # 单独设置背景颜色，无法起作用？
        self.tableWidget.horizontalHeaderItem(2).setForeground(QBrush(Qt.green))

        # 设置主界面的图标
        icon_app = QtGui.QIcon()
        icon_app.addPixmap(QtGui.QPixmap(":/myres/img123/app9.png"))
        self.setWindowIcon(icon_app)
        icon_add_row = QtGui.QIcon()
        icon_add_row.addPixmap(QtGui.QPixmap(":/myres/img123/add9.png"))
        self.addaction.setIcon(icon_add_row)
        icon_delete_row = QtGui.QIcon()
        icon_delete_row.addPixmap(QtGui.QPixmap(":/myres/img123/sub9.png"))
        self.deleteaction.setIcon(icon_delete_row)
        icon_add_item = QtGui.QIcon()
        icon_add_item.addPixmap(QtGui.QPixmap(":/myres/img123/addcir9.png"))
        self.additemaction.setIcon(icon_add_item)
        icon_update_item = QtGui.QIcon()
        icon_update_item.addPixmap(QtGui.QPixmap(":/myres/img123/edit9.png"))
        self.updateitemaction.setIcon(icon_update_item)
        icon_delete_item = QtGui.QIcon()
        icon_delete_item.addPixmap(QtGui.QPixmap(":/myres/img123/subcir9.png"))
        self.deleteitemaction.setIcon(icon_delete_item)
        icon_save_to_db = QtGui.QIcon()
        icon_save_to_db.addPixmap(QtGui.QPixmap(":/myres/img123/save9.png"))
        self.savetodbaction.setIcon(icon_save_to_db)
        icon_excel = QtGui.QIcon()
        icon_excel.addPixmap(QtGui.QPixmap(":/myres/img123/excel9.png"))
        self.exporttoexcelaction.setIcon(icon_excel)
        icon_help = QtGui.QIcon()
        icon_help.addPixmap(QtGui.QPixmap(":/myres/img123/help9.png"))
        self.helpaction.setIcon(icon_help)
        icon_declare = QtGui.QIcon()
        icon_declare.addPixmap(QtGui.QPixmap(":/myres/img123/favor9.png"))
        self.declareaction.setIcon(icon_declare)
        self.qtaction.setIcon(icon_declare)

    def init_db(self):
        if not os.path.exists("peopleinfo.db"):
            # 打开（不存在时候创建）数据库
            mydb = sqlite3.connect("peopleinfo.db")
            # 创建游标
            cu = mydb.cursor()

            # 创建表，暂时没有设置主键
            cu.execute("create table tb_people_info( pid integer primary key," +
                       "name varchar(20), age varchar(3), gender varchar(5)," +
                       "phone varchar(15), info text NULL)")
            cu.close()
            mydb.close()
        else:
            pass

    def add_table_row(self):
        row1 = self.tableWidget.rowCount()
        print(row1)
        self.tableWidget.insertRow(row1)  # 动态增加行的时候，不需要减1

    def delete_table_row(self):
        row2 = self.tableWidget.rowCount()
        if row2 > 1:
            print(row2)
            self.tableWidget.removeRow(row2 - 1)  # 动态删除行的时候，需要减1

    def insert_row(self):  # 添加一行数据，弹出添加用户的dlg
        insertdlg = MyInsertDialog()
        insertdlg.show()
        insertdlg.exec_()

    def update_row(self):  # 更新一行数据
        updatedlg = MyUpdateDialog()
        # 需要在此处添加数据在dlg上面，或者在初始化的时候
        updatedlg.show()
        updatedlg.exec_()

    def delete_row(self):  # 删除一行数据
        # 获取有效的行数
        rowcount = self.tableWidget.rowCount()
        validrow = 0
        for x in range(0, rowcount):
            if (self.tableWidget.item(x, 0) != None):
                validrow = x + 1
            else:
                break
        if (validrow == 0):
            QMessageBox.warning(self, "警告！", "表格之中没有数据！")
            return  # 此处让messagebox位于中间，让dlg不出现
        else:
            index = self.tableWidget.currentRow()
            lineedno = self.tableWidget.item(index, 0)
            lineedname = self.tableWidget.item(index, 1)
            lineedage = self.tableWidget.item(index, 2)
            comboxgender = self.tableWidget.item(index, 3)
            lineedphone = self.tableWidget.item(index, 4)
            lineedotherinfo = self.tableWidget.item(index, 5)
            if (lineedno == None or lineedname == None or lineedage == None or
                        comboxgender == None or lineedphone == None or lineedotherinfo == None):
                QMessageBox.warning(self, "警告！", "所选人员信息不能为空！")
                return  # 此处让messagebox位于中间，让dlg不出现
            else:
                info = "你正在删除第" + str(index + 1) + "行联系人信息，是否确定？"
                reply = QMessageBox.information(self, "删除一行数据", info, QMessageBox.No | QMessageBox.Yes)
                if (reply == QMessageBox.Yes):
                    if (validrow == 1 and index == 0):
                        lineedno.setText("")
                        lineedname.setText("")
                        lineedage.setText("")
                        comboxgender.setText("")
                        lineedphone.setText("")
                        lineedotherinfo.setText("")
                    elif (validrow > 1):
                        self.tableWidget.removeRow(index)
                else:
                    pass

    # 思路:
    # 按照表格的行列号，获得当前有效的行数据，将每一行的数据，存入到一个元组之中，然后再循环插入其中
    # 每一次从插入的时候，检测数据库之中，是否存在已有的数据，对于已有的数据，不去更改，对于新数据，插入，对于删除的数据，删除
    # 现在也可以每次存储之前，都将所有的数据删除，重新插入
    def save_to_db(self):
        try:
            people_info = []
            rowcount = tableuiopt.tableWidget.rowCount()
            # 检测列表之中是否有数据
            if (tableuiopt.tableWidget.item(0, 0) == None or tableuiopt.tableWidget.item(0, 0).text() == ""):
                QMessageBox.warning(self, "提示", "没有可用数据！")
                return

            # 此处还是要修改一下，先处理判断是否为空，如果为空，则不删除，如果不为空，则删除
            mydb = sqlite3.connect("peopleinfo.db")
            cu = mydb.cursor()
            result = cu.execute("SELECT * FROM tb_people_info")
            rowno = len(result.fetchall())
            if rowno > 0:
                cu.execute("DELETE FROM tb_people_info")
                mydb.commit()
            else:
                pass

            for x in range(0, rowcount):
                if (tableuiopt.tableWidget.item(x, 0) != None):
                    pid = tableuiopt.tableWidget.item(x, 0).text()
                    name = tableuiopt.tableWidget.item(x, 1).text()
                    age = tableuiopt.tableWidget.item(x, 2).text()
                    gender = tableuiopt.tableWidget.item(x, 3).text()
                    tel = tableuiopt.tableWidget.item(x, 4).text()
                    other = tableuiopt.tableWidget.item(x, 5).text()
                    temp = (pid, name, age, gender, tel, other)
                    people_info.append(temp)
                    print(people_info)
                else:
                    break
            insertsql = "insert into tb_people_info(pid,name,age,gender,phone,info) VALUES(?,?,?,?,?,?) "
            cu.executemany(insertsql, people_info)
            mydb.commit()
        except FileNotFoundError:
            print("Error: 没有找到数据库文件!")
        except DatabaseError:
            print("DatabaseError: 数据库错误!")
        else:
            print("数据插入成功!")
            cu.close()
            mydb.close()

    def save_to_excel(self):
        # 获取数据
        people_info = []
        rowcount = tableuiopt.tableWidget.rowCount()
        # 检测列表之中是否有数据
        if (tableuiopt.tableWidget.item(0, 0) == None or tableuiopt.tableWidget.item(0, 0).text() == ""):
            QMessageBox.warning(self, "提示", "没有可用数据！")
            return
        else:
            for x in range(0, rowcount):
                if (tableuiopt.tableWidget.item(x, 0) != None):
                    pid = tableuiopt.tableWidget.item(x, 0).text()
                    name = tableuiopt.tableWidget.item(x, 1).text()
                    age = tableuiopt.tableWidget.item(x, 2).text()
                    gender = tableuiopt.tableWidget.item(x, 3).text()
                    tel = tableuiopt.tableWidget.item(x, 4).text()
                    other = tableuiopt.tableWidget.item(x, 5).text()
                    temp = (pid, name, age, gender, tel, other)
                    people_info.append(temp)
                    print(people_info)
                else:
                    break
        try:
            wb = Workbook()  # 创建一个工作簿，一个工作簿创建好之后，默认就有一个sheet
            sheet = wb.active  # 激活当前的sheet
            sheet.title = "联系人信息"
            tableTitle = ['序号', '姓名', '年龄', '性别', '电话', '备注']
            # 如上，openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，
            # 提示 ：Row or column values must be at least 1，即最小值为1.
            # 设置标题
            for col in range(len(tableTitle)):
                c = col + 1
                sheet.cell(row=1, column=c).value = tableTitle[col]
            # 插入数据
            for x in range(len(people_info)):
                for y in range(6):
                    sheet.cell(row=x + 2, column=y + 1).value = people_info[x][y]

            wb.save("peopleinfo.xlsx")  # 保存
        except BaseException:
            print("出现错误！")

    def show_declaration(self):
        QMessageBox.about(self, "关于", "版本:  " + _version + "\n" + "作者:  " + _author)

    def about_qt(self):
        QMessageBox.aboutQt(self, "About Qt")


class MyInsertDialog(QtWidgets.QDialog, InfoDlg.Ui_Dialog):
    __title = "添加联系人信息"

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.second_init()
        self.pushBtnCancle.clicked.connect(self.clear_edit)
        self.pushBtnOkay.clicked.connect(self.insert_row)

    def second_init(self):
        self.setWindowTitle(self.__title)  # 设置了标题栏
        # 将序号屏蔽掉，只能顺序插入，但下面不支持这样做，只能在ui之中操作
        # self.lineEdNo.isEnabled(False)
        self.lineEdNo.setText("人员顺序由系统按顺序自动生成")
        listgender = ("男", "女")
        self.comboBoxGender.addItems(listgender)
        self.setModal(True)  # 设置模态和非模态对话框
        # 也可以使用additem
        # self.comboBoxGender.addItem("男")
        window_icon = QtGui.QIcon()
        window_icon.addPixmap(QtGui.QPixmap(":/myres/img123/addcir9.png"))
        self.setWindowIcon(window_icon)

    def clear_edit(self):
        self.lineEdNo.setText("")
        self.lineEdName.setText("")
        self.lineEdAge.setText("")
        self.lineEdPhone.setText("")
        self.textEdOtherinfo.setText("")

    # 一次只能输入一个cell，如何依次输入多个cell?,可以一次添加多个cell，之前是因为没有找到对应的函数
    # 使得顺序固定增加，不能随意在某行添加了
    def insert_row(self):
        rowcount = tableuiopt.tableWidget.rowCount()
        validrow = 0  # 有效的行数目
        for x in range(0, rowcount):
            if (tableuiopt.tableWidget.item(x, 0) == None or tableuiopt.tableWidget.item(x, 0).text() == ""):
                break
            else:
                validrow = x + 1
        lineedname = self.lineEdName.text()
        lineedage = self.lineEdAge.text()
        comboxgender = self.comboBoxGender.currentText()
        lineedphone = self.lineEdPhone.text()
        lineedotherinfo = self.textEdOtherinfo.toPlainText()  # 获取文本内容
        # 不能使用["|"]运算符，而要使用or运算符，二者的区别，需要注意下
        if (lineedname == '' or lineedage == '' or
                    lineedphone == '' or lineedotherinfo == ''):
            QMessageBox.warning(self, "警告！", "人员信息不能为空！")
            return
        if validrow == 0:
            tableuiopt.tableWidget.setItem(validrow, 0, QTableWidgetItem(str(validrow + 1)))
            tableuiopt.tableWidget.setItem(validrow, 1, QTableWidgetItem(lineedname.strip('')))
            tableuiopt.tableWidget.setItem(validrow, 2, QTableWidgetItem(lineedage.strip('')))
            tableuiopt.tableWidget.setItem(validrow, 3, QTableWidgetItem(comboxgender))
            tableuiopt.tableWidget.setItem(validrow, 4, QTableWidgetItem(lineedphone.strip('')))
            tableuiopt.tableWidget.setItem(validrow, 5, QTableWidgetItem(lineedotherinfo.strip('')))
        else:
            inx = validrow - 1
            indextemp = int(tableuiopt.tableWidget.item(inx, 0).text())  # 用来记录最后一行的ID值
            tableuiopt.tableWidget.setItem(validrow, 0, QTableWidgetItem(str(indextemp + 1)))
            tableuiopt.tableWidget.setItem(validrow, 1, QTableWidgetItem(lineedname.strip('')))
            tableuiopt.tableWidget.setItem(validrow, 2, QTableWidgetItem(lineedage.strip('')))
            tableuiopt.tableWidget.setItem(validrow, 3, QTableWidgetItem(comboxgender))
            tableuiopt.tableWidget.setItem(validrow, 4, QTableWidgetItem(lineedphone.strip('')))
            tableuiopt.tableWidget.setItem(validrow, 5, QTableWidgetItem(lineedotherinfo.strip('')))
        self.close()  # 添加了一个联系人，点击确定之后，自动关闭对话框


class MyUpdateDialog(QtWidgets.QDialog, InfoDlg.Ui_Dialog):
    __title = "更新联系人信息"

    def __init__(self, parent=None):
        super(MyUpdateDialog, self).__init__(parent)
        self.setupUi(self)
        self.second_init()
        self.pushBtnCancle.clicked.connect(self.clear_edit)
        self.pushBtnOkay.clicked.connect(self.update_row)

    def second_init(self):
        self.setWindowTitle(self.__title)
        listgender = ("男", "女")
        self.comboBoxGender.addItems(listgender)
        window_icon = QtGui.QIcon()
        window_icon.addPixmap(QtGui.QPixmap(":/myres/img123/edit9.png"))
        self.setWindowIcon(window_icon)

        # 在此处给修改的dlg添加数据
        # print(tableuiopt.tableWidget.item(rowindex, 0).text())  # 获取到了某一个表格的内容
        rowindex = 0  # 获取行号
        if (tableuiopt.tableWidget.currentRow() == 0):
            rowindex = 0
        else:
            rowindex = tableuiopt.tableWidget.currentRow()
        # 此处可以直接使用tableuiopt.tableWidget.item(rowindex, 0).text(),
        # 但是如果遇到了text内容为空的情况，就无法处理了，会报错，所以使用None来先验证
        lineedno = tableuiopt.tableWidget.item(rowindex, 0)
        lineedname = tableuiopt.tableWidget.item(rowindex, 1)
        lineedage = tableuiopt.tableWidget.item(rowindex, 2)
        comboxgender = tableuiopt.tableWidget.item(rowindex, 3)
        lineedphone = tableuiopt.tableWidget.item(rowindex, 4)
        lineedotherinfo = tableuiopt.tableWidget.item(rowindex, 5)
        if (lineedno == None or lineedname == None or lineedage == None or
                    comboxgender == None or lineedphone == None or lineedotherinfo == None):
            # s = QMessageBox()
            # s.setGeometry(self, 200, 300)
            QMessageBox.warning(self, "警告！", "所选人员信息不能为空！")
            return  # 此处让messagebox位于中间，让dlg不出现
        else:
            self.lineEdNo.setText(lineedno.text())
            self.lineEdName.setText(lineedname.text())
            self.lineEdAge.setText(lineedage.text())
            self.comboBoxGender.setCurrentText(comboxgender.text())
            self.lineEdPhone.setText(lineedphone.text())
            self.textEdOtherinfo.setText(lineedotherinfo.text())  # 获取文本内容

    def clear_edit(self):
        self.lineEdNo.setText("")
        self.lineEdName.setText("")
        self.lineEdAge.setText("")
        self.lineEdPhone.setText("")
        self.textEdOtherinfo.setText("")

    def update_row(self):
        lineedno = self.lineEdNo.text()
        lineedname = self.lineEdName.text()
        lineedage = self.lineEdAge.text()
        comboxgender = self.comboBoxGender.currentText()
        lineedphone = self.lineEdPhone.text()
        lineedotherinfo = self.textEdOtherinfo.toPlainText()  # 获取文本内容
        # 不能使用["|"]运算符，而要使用or运算符，二者的区别，需要注意下
        if (lineedno == '' or lineedname == '' or lineedage == '' or
                    lineedphone == '' or lineedotherinfo == ''):
            print("111")
            QMessageBox.warning(self, "警告！", "人员信息不能为空！")
            return 0
        rowindex = 0
        if tableuiopt.tableWidget.selectedItems() is None:
            rowindex = 0
        else:
            rowindex = tableuiopt.tableWidget.currentRow()
            print(rowindex)
            tableuiopt.tableWidget.setItem(rowindex, 0, QTableWidgetItem(lineedno.strip('')))
            tableuiopt.tableWidget.setItem(rowindex, 1, QTableWidgetItem(lineedname.strip('')))
            tableuiopt.tableWidget.setItem(rowindex, 2, QTableWidgetItem(lineedage.strip('')))
            tableuiopt.tableWidget.setItem(rowindex, 3, QTableWidgetItem(comboxgender))
            tableuiopt.tableWidget.setItem(rowindex, 4, QTableWidgetItem(lineedphone.strip('')))
            tableuiopt.tableWidget.setItem(rowindex, 5, QTableWidgetItem(lineedotherinfo.strip('')))
        self.close()  # 更新了一个联系人，点击确定之后，自动关闭对话框


if __name__ == "__main__":
    app = QApplication(sys.argv)
    tableuiopt = TableUiOpt()
    # tableuiopt.inittable() #此处初始化也可以，但是不太规范
    tableuiopt.show()
    sys.exit(app.exec_())
